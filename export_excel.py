import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def export_to_excel(df, enterprise_name, date_str):
    # Filename with timestamp
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tva_report_{enterprise_name}_{current_time}.xlsx"
    output_path = os.path.join("generated_reports", filename)
    os.makedirs("generated_reports", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "TVA Report"

    # Set row height for all rows (e.g., 22 instead of 30)
    for i in range(1, 100):  # Adjust 100 to max expected rows
        ws.row_dimensions[i].height = 22

    # Set column width for relevant columns (e.g., 18 instead of 30)
    for col in range(2, 7):  # Columns B to G (your table columns)
        ws.column_dimensions[chr(64 + col)].width = 18

    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    grey_fill = PatternFill(start_color='C0C0C0',
                            end_color='C0C0C0', fill_type='solid')
    green_fill = PatternFill(start_color="C6EFCE",
                             end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE",
                           end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C",
                              end_color="FFEB9C", fill_type="solid")

    def write_section(start_row, title, data, is_client, first_col_header):
        # Merge and set section title
        ws.merge_cells(start_row=start_row, start_column=2,
                       end_row=start_row, end_column=6)
        ws.cell(row=start_row, column=2).value = title
        ws.cell(row=start_row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=start_row, column=2).alignment = center

        headers = [first_col_header, "MT TTC", "M. H.T", "Taux TVA", "TVA"]
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=start_row + 1, column=col_num)
            cell.value = header
            cell.font = bold
            cell.alignment = center
            cell.border = border

        tva_total = 0
        for i, row in enumerate(data, start=start_row + 2):
            for j, key in enumerate(["Service", "TTC", "HT", "TVA Rate", "TVA"], 2):
                cell = ws.cell(row=i, column=j)
                if key == "TVA Rate":
                    cell.value = f"{row[key]}%"
                    cell.fill = grey_fill
                else:
                    cell.value = row[key]
                cell.border = border
                # Make text bold and moderately bigger
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                if key == "TVA":
                    tva_total += row[key]
                if row['Service'] == "Crédit Précédent":
                    cell.fill = red_fill
                elif row['Service'].upper().startswith("FAC"):
                    cell.fill = yellow_fill
                elif row['Service'].upper().startswith("FACTURE"):
                    cell.fill = red_fill

        total_row = start_row + 2 + len(data)
        ws.merge_cells(start_row=total_row, start_column=2,
                       end_row=total_row, end_column=5)
        ws.cell(row=total_row, column=2).value = "la somme de TVA"
        ws.cell(row=total_row, column=2).fill = green_fill
        ws.cell(row=total_row, column=2).font = bold
        ws.cell(row=total_row, column=2).alignment = center
        ws.cell(row=total_row, column=6).value = round(tva_total, 2)
        ws.cell(row=total_row, column=6).font = bold
        ws.cell(row=total_row, column=6).fill = green_fill

        return tva_total, total_row

    # Filter roles
    clients = df[df['Role'] == 'Client'].to_dict(orient='records')
    fournisseurs = df[df['Role'] == 'Fournisseur'].to_dict(orient='records')

    # Move "Crédit Précédent" entries to the end of fournisseurs
    credit_precedent_entries = [
        f for f in fournisseurs if f.get("Service") == "Crédit Précédent"]
    other_fournisseurs = [f for f in fournisseurs if f.get(
        "Service") != "Crédit Précédent"]
    fournisseurs = other_fournisseurs + credit_precedent_entries

    # --- Write CLIENTS first at the top ---
    start_row_clients = 3
    ca_title = f"C.A du {date_str}  {enterprise_name}"
    tva_client, end_row_clients = write_section(
        start_row_clients, ca_title, clients, is_client=True, first_col_header="Ventes")

    # Nombre de facture (clients only)
    ws.cell(row=end_row_clients + 1, column=2).value = "Nombre de Facture"
    ws.cell(row=end_row_clients + 1, column=3).value = len(clients)
    ws.cell(row=end_row_clients + 1, column=2).font = bold
    ws.cell(row=end_row_clients + 1, column=2).fill = green_fill
    ws.cell(row=end_row_clients + 1, column=3).font = bold
    ws.cell(row=end_row_clients + 1, column=3).fill = green_fill

    # --- Write FOURNISSEURS after clients ---
    start_row_fournisseurs = end_row_clients + 3
    tva_title = f"TVA RECUPERABLE le {date_str} "
    tva_fournisseur, end_row_fournisseurs = write_section(
        start_row_fournisseurs, tva_title, fournisseurs, is_client=False, first_col_header="Achats")

    # TVA à payer at the bottom
    final_row = end_row_fournisseurs + 3
    ws.merge_cells(start_row=final_row, start_column=2,
                   end_row=final_row, end_column=5)
    ws.cell(row=final_row, column=2).value = "TVA DUE"
    ws.cell(row=final_row, column=2).font = Font(bold=True, color="FF0000")
    ws.cell(row=final_row, column=2).fill = red_fill
    ws.cell(row=final_row, column=2).alignment = center
    ws.cell(row=final_row, column=6).value = round(
        tva_client - tva_fournisseur, 2)
    ws.cell(row=final_row, column=6).font = Font(bold=True)

    # Save file
    wb.save(output_path)
    return output_path
